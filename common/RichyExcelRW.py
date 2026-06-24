import time
import openpyxl
from ddt import ddt,data,unpack
from shutil import copyfile
import sys
import os
FILENAME=""

def rinit(filename):
    print("init",filename)
     #获取路径中文件名
    global FILENAME
    if FILENAME =="":
        file = os.path.basename(filename)
        newfilename=os.path.join(os.path.dirname(filename), file.split(".")[0] + time.strftime('%Y%m%d%H%M%S') + "." + file.split(".")[1])
        rfilecopy(filename,newfilename)
        print("copyfile")
        FILENAME=newfilename
    return FILENAME
def rread(SHEETNAME):
    #读取excel表格
    global FILENAME
    # global SHEETNAME
    print("FILENAME",FILENAME)
    workbook=openpyxl.load_workbook(FILENAME)
    #获取表内容
    sheet=workbook[SHEETNAME]
    rows=list(sheet.rows)
    #获取表头
    title = [i.value for i in rows[0]]
    #tile=['id','title','param','assert']

    #获取参数化数据
    cases=[]
    for case in rows[1:]:
        case_data=list(map(lambda x:"" if x.value is None else x.value,case))
        print(case_data)
        case_dict=dict(zip(title,case_data))
        cases.append(case_dict)
    #[{'username':'test1','pwd':'123456',ele},{'username':'tes1t','pwd':'123456',ele},{}]
    #返回读取的所有数据
    print(cases)
    workbook.close()
    return cases
# def rwrite(SHEETNAME,row,column,value):
def rwrite(SHEETNAME,row,column,value):
    global FILENAME

    print("write",FILENAME,SHEETNAME)
    workbook=openpyxl.load_workbook(FILENAME)
    sheet = workbook[SHEETNAME]
    sheet.cell(row=row,column=column,value=value)
    workbook.save(FILENAME)
    workbook.close()
def rfilecopy(source,target):
    try:
        print(source,target)
        copyfile(source, target)
    except IOError as e:
        print("Unable to copy file. %s" % e)
    except:
        print("Unexpected error:", sys.exc_info())
def rdata(SHEETNAME):
    return ridata(rread(SHEETNAME))
def ridata(iterable, index_len=None):
    if index_len is None:
        iterable = tuple(iterable)
        index_len = len(str(len(iterable)))
    def wrapper(func):
        setattr(func, '%values', iterable)
        setattr(func, '%index_len', index_len)
        return func
    return wrapper


if __name__ == "__main__":
    rinit('../datas/cloudtranslation.xlsx')
    print(rread("login"))
    # excelrw.write(2,5,'成功')

